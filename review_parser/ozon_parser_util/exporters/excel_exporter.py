"""Excel (.xlsx) exporter for Ozon reviews."""
from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ozon_parser_util.core.models import OzonReview
from ozon_parser_util.exporters.base import BaseExporter

logger = logging.getLogger(__name__)

_HEADER_BG    = "005BFF"   # Ozon blue
_HEADER_FONT  = "FFFFFF"
_ALT_ROW_BG   = "EEF3FF"
_MAX_COL_WIDTH = 70


class ExcelExporter(BaseExporter):
    def export(self, reviews: list[OzonReview], output_path: Path) -> None:
        self._ensure_parent(output_path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отзывы Ozon"

        headers     = OzonReview.column_names()
        field_names = list(OzonReview.__dataclass_fields__.keys())

        header_font  = Font(bold=True, color=_HEADER_FONT, size=11)
        header_fill  = PatternFill("solid", fgColor=_HEADER_BG)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
        ws.row_dimensions[1].height = 22

        alt_fill = PatternFill("solid", fgColor=_ALT_ROW_BG)
        wrap     = Alignment(vertical="top", wrap_text=True)

        for row_idx, review in enumerate(reviews, start=2):
            data = review.to_dict()
            for col_idx, field in enumerate(field_names, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=data[field])
                cell.alignment = wrap
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            width = min(max_len + 3, _MAX_COL_WIDTH)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width

        ws.freeze_panes = "A2"
        wb.save(str(output_path))
        logger.info("Excel saved → %s  (%d rows)", output_path, len(reviews))