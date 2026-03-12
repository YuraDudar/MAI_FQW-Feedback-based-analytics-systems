"""
Excel (.xlsx) exporter for WB reviews.

Layout:
  Sheet "Отзывы":
    - Main review table with Russian column headers.
    - Data Dictionary sidebar (2-column gap to the right of main data).

Formatting:
  - NaN values → empty cell.
  - Alternating row shading for readability.
  - Frozen header row + auto-width columns.
  - Data Dictionary colour-coded by group (steel-blue header).
"""
from __future__ import annotations

import json
import logging
import math
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from wb_parser_util.core.models import DATA_DICT, Review
from wb_parser_util.exporters.base import BaseExporter

logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
_MAIN_HDR_BG   = "5B2D8E"
_MAIN_HDR_FG   = "FFFFFF"
_ALT_ROW_BG    = "F3EEF9"
_DICT_HDR_BG   = "2E6DA4"
_DICT_HDR_FG   = "FFFFFF"
_MAX_COL_WIDTH = 60
_DICT_GAP_COLS = 2

_GROUP_COLOURS: dict[str, str] = {
    "Метаданные":       "FFE6CC",
    "Идентификация":    "CCE5FF",
    "Автор":            "D4EDDA",
    "Контент отзыва":   "FFF3CD",
    "Даты и статус":    "F8D7DA",
    "Ответ продавца":   "E2CFEE",
    "Соответствия":     "D1ECF1",
    "Голоса и рейтинг": "FFEEBA",
    "Медиа":            "C3E6CB",
    "Исключение":       "F5C6CB",
    "Причины оценки":   "BEE5EB",
}


class ExcelExporter(BaseExporter):
    """Writes reviews into a formatted .xlsx file with a Data Dictionary sidebar."""

    def export(self, reviews: list[Review], output_path: Path) -> None:
        self._ensure_parent(output_path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отзывы"

        field_names = Review.field_names()
        ru_headers  = Review.column_names_ru()
        n_cols      = len(field_names)

        _write_main_header(ws, ru_headers)
        _write_main_data(ws, reviews, field_names)
        _autofit_columns(ws, n_cols)
        ws.freeze_panes = "A2"

        dict_start_col = n_cols + 1 + _DICT_GAP_COLS
        _write_data_dictionary(ws, dict_start_col)

        wb.save(str(output_path))
        logger.info("Excel saved → %s  (%d rows)", output_path, len(reviews))


# ── Main table helpers ────────────────────────────────────────────────────────

def _write_main_header(ws, headers: list[str]) -> None:
    hdr_font  = Font(bold=True, color=_MAIN_HDR_FG, size=11)
    hdr_fill  = PatternFill("solid", fgColor=_MAIN_HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
    ws.row_dimensions[1].height = 28


def _write_main_data(ws, reviews: list[Review], field_names: list[str]) -> None:
    alt_fill = PatternFill("solid", fgColor=_ALT_ROW_BG)
    wrap     = Alignment(vertical="top", wrap_text=True)

    for row_idx, review in enumerate(reviews, start=2):
        data = review.to_dict()
        for col_idx, field in enumerate(field_names, start=1):
            cell_val = _to_cell(data[field])
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
            cell.alignment = wrap
            if row_idx % 2 == 0:
                cell.fill = alt_fill


def _autofit_columns(ws, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, _MAX_COL_WIDTH)


# ── Data Dictionary sidebar ───────────────────────────────────────────────────

_DICT_COLS   = ["Поле (RU)", "Поле (EN)", "Тип", "Группа", "Описание"]
_DICT_WIDTHS = [28, 26, 12, 20, 55]


def _write_data_dictionary(ws, start_col: int) -> None:
    hdr_font  = Font(bold=True, color=_DICT_HDR_FG, size=11)
    hdr_fill  = PatternFill("solid", fgColor=_DICT_HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, title in enumerate(_DICT_COLS):
        c = start_col + i
        cell = ws.cell(row=1, column=c, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    for i, width in enumerate(_DICT_WIDTHS):
        ws.column_dimensions[get_column_letter(start_col + i)].width = width

    wrap = Alignment(vertical="top", wrap_text=True)

    for row_offset, (field, ru_name, typ, group, desc) in enumerate(DATA_DICT, start=2):
        group_bg = _GROUP_COLOURS.get(group, "EBF2FA")
        row_fill = PatternFill("solid", fgColor=group_bg)

        for i, val in enumerate([ru_name, field, typ, group, desc]):
            c = start_col + i
            cell = ws.cell(row=row_offset, column=c, value=val)
            cell.fill      = row_fill
            cell.alignment = wrap
            if i == 1:
                cell.font = Font(bold=True, size=10)

    _apply_outer_border(ws, 1, start_col, len(DATA_DICT) + 1, start_col + len(_DICT_COLS) - 1)


def _apply_outer_border(ws, r1, c1, r2, c2) -> None:
    thin = Side(style="thin", color="888888")
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = Border(
                top    = thin if cell.row    == r1 else None,
                bottom = thin if cell.row    == r2 else None,
                left   = thin if cell.column == c1 else None,
                right  = thin if cell.column == c2 else None,
            )


# ── Value helper ──────────────────────────────────────────────────────────────

def _to_cell(value: Any) -> Any:
    """NaN → None (empty cell); list/dict → JSON string; else as-is."""
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value